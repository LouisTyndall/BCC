# Import `load_workbook` module from `openpyxl`
from openpyxl import load_workbook
import os
from collections import defaultdict
d='counts'
m=1.43
files=os.listdir(d)
# Load in the workbook
data=defaultdict(float)
for f in files:
	wb = load_workbook(d+'/'+f)
	print (f)
	# Get sheet names

	sheet=wb[wb.sheetnames[0]]

	for n in range(2,sheet.max_row+1):
		if sheet['D'+str(n)].value==None:
			data[sheet['A'+str(n)].value,sheet['B'+str(n)].value]+=float(sheet['C'+str(n)].value)
		else:
			data[sheet['A'+str(n)].value,sheet['B'+str(n)].value]+=float(sheet['D'+str(n)].value)
print (data)
j={n:int(data[n]*m) for n in data}
print (j)
import pickle

with open('AADT.json','wb') as f:
	pickle.dump(j,f)
	
